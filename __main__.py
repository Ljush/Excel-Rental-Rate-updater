from openpyxl import load_workbook
from helpers.verify_OS import current_os
from helpers.check_date import find_current_month
from helpers.timer import Timer
#from helpers.dir_fix import *
#from tkinter import Tk, filedialog

from southwest_updater import *
from leduc_updater import *
from west_updater import *
from millwoods_updater import *
from downtown_updater import *
from sherwood_park_updater import *
from north_east_updater import *
from grande_prairie_updater import *
'''Openpyxl documentation
https://openpyxl.readthedocs.io/en/stable/
https://openpyxl.readthedocs.io/en/latest/tutorial.html#loading-from-a-file'''


def fix_mac_directory():
    """Prompt the user with a window to select a folder as a directory.
        The user should be selecting the 'sheets' folder in the project
        as the desired directory. This is where all resulting excel sheets will
        be saved in the future."""
    root = Tk()
    root.withdraw()
    foldr = filedialog.askdirectory()
    with open('mac_directories.txt', 'w') as f:
        print(f); print(foldr)
        f = f.write((str(foldr)))
        print('3')
    print('ssssss')
    print('foldr:', foldr)
    print(f)
    #f.close()
    print('4')
    sheets_folder = foldr
    return sheets_folder

def fix_win_directory():
    """Prompt the user with a window to select a folder as a directory.
        The user should be selecting the 'sheets' folder in the project
        as the desired directory. This is where all resulting excel sheets will
        be saved in the future."""
    root = Tk()
    root.withdraw()
    foldr = filedialog.askdirectory()
    with open('win_directories.txt', 'w') as file_to_write:
        write_to_folder = file_to_write.write((str(foldr)))
    print(write_to_folder)
    print('done, closing now.')
    write_to_folder.close()
    sheets_folder = foldr
    return sheets_folder

def verify_directories():
    """holder"""
    pass

def get_excel_path():
    """Prompt the user to select the excel file to update/scrape date onto."""
    root = Tk()
    root.withdraw()
    file = filedialog.askopenfilename()
    return file

def main():
    main_timer = Timer()
    current_system = current_os()
    if current_system == 'windows':
        with open('windows_directories.txt', 'r') as write_dir:
            sheets_folder = write_dir.read()
     
    if current_system == 'darwin':
        # open directory txt file to get string of user pathway to sheets folder
        try:
            with open('mac_directories.txt', 'r') as write_dir:
                sheets_folder = write_dir.read()
        # otherwise create the txt file and continue as normal
        except:
            new = open('mac_directories.txt', 'x')
            sheets_folder = new.read()

    # when the directory txt file is empty; 
    if sheets_folder == '':
        print('*****PLEASE SELECT THE SHEETS FOLDER*****')
        sheets_folder = fix_mac_directory()
        print('\n****SELECT COMPETITION EXCEL FILE*****\n')
        open_file = get_excel_path()
        rates = load_workbook(open_file)
                
    # when str directory is found
    if sheets_folder != '':
        # prompt user for the excel sheet to add on-to.
        open_file = get_excel_path()
        rates = load_workbook(open_file)
    return
    # go left to right on individual sheets for spreadsheet; timer posts at end total runtime in seconds.
    leduc = rates['Leduc']
    main_timer.start()
    leduc_updater(leduc)
    print('\nLeduc sheet done.')

    west = rates['West']
    west_updater(west)
    print('\nWest sheet done.')

    millwoods = rates['Millwoods']
    millwoods_updater(millwoods)
    print('\nMillwoods sheet done.')
    
    sw = rates['Southwest']
    southwest_updater(sw)
    print('\nSouthwest sheet done.')
    
    downtown = rates['Downtown']
    downtown_updater(downtown)
    print('\nDowntown sheet done.')

    ne = rates['North East']
    north_east_updater(ne)
    print('\nNorth East sheet done.')

    sherwood = rates['Sherwood Park']
    sherwood_updater(sherwood)
    print('\nSherwood Park sheet done.')

    gp = rates['Grande Prairie']
    grande_prairie(gp)
    print('\nGrande Prairie sheet done.\n')
    main_timer.stop()

    day_int, month_int, year_int, month_dict, curr_time = find_current_month()
    # Close and save the excel sheet xlsx
    sheets = f'{sheets_folder}/sheets/'
    try:
        if current_system == 'windows':
            rates.save(r'{}\{} {} - Competitionb Rates - AB.xlsx'.format(
                sheets_folder, month_dict[month_int], year_int))
        if current_system == 'darwin':
            rates.save(
                '{}{} {} - Competition Rates - AB.xlsx'.format(sheets, month_dict[month_int], year_int))
        print('\nCompetition Rate excel sheet saved.\n')
    except:
        raise OSError('Could not save file. Check your file/save directories.')

if __name__ == '__main__':
    main()